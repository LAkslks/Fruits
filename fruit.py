import openpyxl
import json


with open('color.json', 'r', encoding='utf-8') as f:
    data = json.load(f)
    

    
book = openpyxl.load_workbook('my_book.xlsx')

sheet = book.active

sheet['A1'] = 'fruit'
sheet['B1'] = 'size'
sheet['C1'] = 'color'


row = 2

for item in data:
    sheet.cell(row=row, column= 1, value=item['fruit'])
    sheet.cell(row=row, column= 2, value=item['size'])
    sheet.cell(row=row, column= 3, value=item['color'])
    row += 1
book.save('my_book.xlsx')


data = []   

for row in range(1, 5 +1):
    
    item = ({
        "fruit": str(sheet.cell(row=row, column= 1,).value),
        "size": str(sheet.cell(row=row, column= 2,).value),
        "color":str(sheet.cell(row=row, column= 3,).value)
    })
    
    data.append(item)
    
